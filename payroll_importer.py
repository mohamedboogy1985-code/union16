import re
from openpyxl import load_workbook
from db import get_connection


def _norm(text):
    text = str(text or "").replace("ـ", "")
    text = re.sub(r"\s+", "", text)
    text = (text.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
                .replace("ى", "ي").replace("ة", "ه"))
    return text


def _clean(value):
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def _find_columns(ws):
    """تحديد مواقع الأعمدة من الترويسات (الصفوف 3-5) بشكل ديناميكي،
    مع الاحتفاظ بالأعمدة المعروفة كقيم احتياطية إن تعذر الاكتشاف."""
    fallback = {
        "employee_no": 1, "name": 2,
        "gross": 18, "deductions": 30, "net": 31,
        "salary_advance": 24, "bank_loan": 25,
        "insurance": 22, "tax": 29,
    }
    cols = {}
    for r in range(3, 6):
        for c in range(1, ws.max_column + 1):
            h = _norm(ws.cell(r, c).value)
            if not h:
                continue
            if c not in cols:
                if h == "الاسم":
                    cols["name"] = c
                elif _norm("الجمله") == h and c not in cols.values():
                    # الجملة (الإجمالي) فقط وليست "جملة المستقطع"
                    if "مستقطع" not in h:
                        cols["gross"] = c
                if "المستقطع" in h:
                    cols["deductions"] = c
                if h.startswith("صافي") or "المستحق" in h:
                    cols["net"] = c
                if "قسط" in h and "سلفة" in h and "بنك" not in h:
                    cols["salary_advance"] = c
                if "سلفة" in h and "بنك" in h:
                    cols["bank_loan"] = c
                if "اشتراك" in h and "تامين" in h:
                    cols["insurance"] = c
                if "ضرائب" in h:
                    cols["tax"] = c
                if h == "م" and c == 1:
                    cols["employee_no"] = c
    for key, value in fallback.items():
        cols.setdefault(key, value)
    return cols


def _num(value):
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def import_payroll_excel(file_path: str) -> str:
    wb = load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    title = _clean(ws.cell(2, 1).value) or _clean(ws.title)

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM payroll_rows")
    cur.execute("DELETE FROM payroll_imports")
    conn.commit()

    cols = _find_columns(ws)

    # تحقق: إذا كانت القيم المحسوبة (من صيغ) غير مخزنة في الملف سنستخدم صيغها.
    probe = ws.cell(6, cols["gross"]).value
    if probe is None:
        wb2 = load_workbook(file_path, data_only=False)
        ws2 = wb2[wb2.sheetnames[0]]
        proto = _clean(ws2.cell(6, cols["gross"]).value)
        if proto.startswith("=") or "=" in proto:
            raise ValueError(
                "الملف يحتوي صيغاً (مثل =SUM) دون حفظ ناتجها. "
                "افتح الملف في Excel ثم احفظه (File > Save) وأعد المحاولة."
            )

    rows_count = 0
    total_gross = 0.0
    total_deductions = 0.0
    total_net = 0.0
    total_salary_advance = 0.0
    total_bank_loan = 0.0

    imported_at = __import__("datetime").datetime.now().isoformat(timespec="seconds")
    cur.execute(
        "INSERT INTO payroll_imports(source_file, imported_at, payroll_month) VALUES (?, ?, ?)",
        (file_path, imported_at, title)
    )
    import_id = cur.lastrowid

    for row in range(6, ws.max_row + 1):
        name = _clean(ws.cell(row, cols["name"]).value)
        if not name:
            continue
        normalized_name = name.replace("ـ", "").strip()
        if "الاجمال" in normalized_name or "الإجمال" in normalized_name:
            break

        employee_no = ws.cell(row, cols["employee_no"]).value
        if employee_no in (None, ""):
            continue
        gross = _num(ws.cell(row, cols["gross"]).value)
        deductions = _num(ws.cell(row, cols["deductions"]).value)
        net = _num(ws.cell(row, cols["net"]).value)
        salary_advance = _num(ws.cell(row, cols["salary_advance"]).value)
        bank_loan = _num(ws.cell(row, cols["bank_loan"]).value)
        insurance = _num(ws.cell(row, cols["insurance"]).value)
        tax = _num(ws.cell(row, cols["tax"]).value)

        cur.execute("""
            INSERT INTO payroll_rows(
                import_id, employee_no, employee_name, payroll_month,
                gross_total, total_deductions, net_pay,
                salary_advance_installment, bank_loan_installment,
                insurance_employee, tax_amount
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            import_id, int(employee_no) if employee_no not in (None, "") else None, name, title,
            gross, deductions, net, salary_advance, bank_loan, insurance, tax
        ))

        rows_count += 1
        total_gross += gross
        total_deductions += deductions
        total_net += net
        total_salary_advance += salary_advance
        total_bank_loan += bank_loan

    cur.execute("""
        UPDATE payroll_imports
        SET employees_count=?, gross_total=?, deductions_total=?, net_total=?,
            salary_advance_total=?, bank_loan_total=?
        WHERE id=?
    """, (rows_count, total_gross, total_deductions, total_net, total_salary_advance, total_bank_loan, import_id))
    conn.commit()
    conn.close()

    return (
        f"تم استيراد ملف المرتبات بنجاح\n"
        f"الفترة: {title}\n"
        f"عدد العاملين: {rows_count}\n"
        f"إجمالي الجملة: {total_gross:,.2f}\n"
        f"إجمالي المستقطعات: {total_deductions:,.2f}\n"
        f"صافي المستحق: {total_net:,.2f}\n"
        f"إجمالي قسط السلف: {total_salary_advance:,.2f}\n"
        f"إجمالي قسط سلفة البنك: {total_bank_loan:,.2f}"
    )