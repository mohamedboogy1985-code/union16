
import os
import re
from datetime import datetime
from openpyxl import load_workbook
from db import init_db, get_connection


def _norm_ar(text):
    """توحيد أشكال الإملاء العربي لتقارن الترويسات والعناوين بشكل موثوق."""
    text = str(text or "").replace("ـ", "")
    text = re.sub(r"\s+", "", text)
    text = (text.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
                .replace("ى", "ي").replace("ة", "ه"))
    return text


def _is_total_label(text):
    return "الاجمال" in _norm_ar(text)

def guess_category(account_name: str) -> str:
    name = account_name.strip()
    if "بنك" in name or "خزينة" in name or "نقد" in name:
        return "نقدية"
    if "إيراد" in name:
        return "إيرادات"
    if "مصروف" in name or "بدل" in name or "اتعاب" in name or "اعانات" in name or "إعانات" in name:
        return "مصروفات"
    if "مجمع اهلاك" in name:
        return "إهلاك"
    if "مدينون" in name:
        return "مدينون"
    if "دائنون" in name:
        return "دائنون"
    if "اصول" in name or "أصول" in name or "سيارات" in name or "مبانى" in name or "مباني" in name:
        return "أصول"
    if "قروض" in name or "احتياطي" in name or "الفائض" in name:
        return "حقوق ملكية / التزامات"
    return "أخرى"

def guess_normal_side(category: str) -> str:
    if category in ("إيرادات", "دائنون", "حقوق ملكية / التزامات"):
        return "credit"
    return "debit"

def _clean_name(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)) and value == 0:
        return ""
    return str(value).replace("\n", " ").strip()

def _num(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


# أسماء الأعمدة المعروفة لمراكز التكلفة داخل ملف اليومية
_COST_CENTER_HEADERS = {
    "مركز التكلفة", "مركز تكلفة", "مركز التكاليف", "مراكز التكلفة",
    "cost center", "cost_center", "costcentre",
}


def _normalize_header(value):
    """تطبيع نص عنوان العمود للمقارنة (يزيل المسافات والفواصل)."""
    s = str(value or "").replace("\n", " ").strip().lower()
    s = s.replace("_", " ").replace("-", " ").replace("ـ", "")
    return " ".join(s.split())


def _find_cost_center_column(ws):
    """يبحث في صفوف العناوين عن عمود مراكز التكلفة ويعيد رقم العمود أو None."""
    for row_idx in (2, 3):
        for col in range(1, ws.max_column + 1):
            header = _normalize_header(ws.cell(row_idx, col).value)
            if header in _COST_CENTER_HEADERS:
                return col
    return None


def _clean_center_name(value):
    """تنظيف اسم مركز التكلفة من الخلية."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return "" if value == 0 else str(value).strip()
    s = str(value).replace("\n", " ").replace("ـ", "").strip()
    return " ".join(s.split())


def import_excel_file(file_path: str) -> str:
    init_db()
    wb = load_workbook(file_path, data_only=False)
    ws = wb[wb.sheetnames[0]]

    # 1) كشف خانتي الإجمالى من الترويسة ثم تجميع أزواج أعمدة الحسابات (مدين/دائن)
    totals = None
    candidates = []
    for col in range(4, ws.max_column + 1, 2):
        raw_name = _clean_name(ws.cell(2, col).value)
        if _is_total_label(raw_name):
            if totals is None:
                totals = (col, col + 1)
            continue
        if raw_name:
            candidates.append((raw_name, col, col + 1))

    # عمودا الإجمالى وفق تخطيط اليومية الأمريكية هما 4 و 5 عادة.
    if totals is None:
        totals = (4, 5)

    account_pairs = []
    seen = {}
    for raw_name, debit_col, credit_col in candidates:
        if debit_col in totals or credit_col in totals:
            continue
        count = seen.get(raw_name, 0) + 1
        seen[raw_name] = count
        final_name = raw_name if count == 1 else f"{raw_name} ({count})"
        account_pairs.append((final_name, debit_col, credit_col))

    cost_center_col = _find_cost_center_column(ws)

    conn = get_connection()
    cur = conn.cursor()

    # clear with same connection
    cur.execute("DELETE FROM journal_lines")
    cur.execute("DELETE FROM journal_entries")
    cur.execute("DELETE FROM accounts")
    conn.commit()

    account_id_map = {}
    for i, (name, _, _) in enumerate(account_pairs, start=1):
        category = guess_category(name)
        normal_side = guess_normal_side(category)
        code = f"A{i:04d}"
        cur.execute(
            "INSERT INTO accounts(code, name, category, normal_side) VALUES (?, ?, ?, ?)",
            (code, name, category, normal_side)
        )
        account_id_map[name] = cur.lastrowid

    entries_count = 0
    lines_count = 0
    centers_created = []
    center_ids = {}

    def ensure_cost_center(name):
        """يستخرج مركز تكلفة من الاسم (ينشئه إن لم يكن موجوداً) ويعيد id أو None."""
        if name in center_ids:
            return center_ids[name]
        row = cur.execute("SELECT id FROM cost_centers WHERE name = ?", (name,)).fetchone()
        if row:
            center_ids[name] = row[0]
            return row[0]
        now = datetime.now().isoformat(timespec="seconds")
        cur.execute(
            "INSERT INTO cost_centers(name, description, created_at) VALUES (?, ?, ?)",
            (name, "مستورد تلقائياً من ملف اليومية", now)
        )
        center_ids[name] = cur.lastrowid
        centers_created.append(name)
        return cur.lastrowid

    for row in range(4, ws.max_row + 1):
        date_value = ws.cell(row, 1).value
        description = _clean_name(ws.cell(row, 3).value)
        reference = _clean_name(ws.cell(row, 2).value)
        total_debit = _num(ws.cell(row, totals[0]).value)
        total_credit = _num(ws.cell(row, totals[1]).value)

        if _is_total_label(date_value):
            break
        if not date_value and not description:
            continue

        if hasattr(date_value, "strftime"):
            entry_date = date_value.strftime("%Y-%m-%d")
        else:
            entry_date = str(date_value)

        # مركز تكلفة القيد كاملاً من العمود المخصص (إن وجد)
        entry_cost_id = None
        if cost_center_col:
            center_name = _clean_center_name(ws.cell(row, cost_center_col).value)
            if center_name and center_name not in ("الإجمالى", "الإجمـــالى"):
                entry_cost_id = ensure_cost_center(center_name)

        cur.execute(
            """
            INSERT INTO journal_entries(entry_date, reference, description, total_debit, total_credit, source_row)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (entry_date, reference, description, total_debit, total_credit, row)
        )
        entry_id = cur.lastrowid
        line_added = 0

        for account_name, debit_col, credit_col in account_pairs:
            debit = _num(ws.cell(row, debit_col).value)
            credit = _num(ws.cell(row, credit_col).value)
            if abs(debit) < 1e-9 and abs(credit) < 1e-9:
                continue
            cur.execute(
                """
                INSERT INTO journal_lines(entry_id, account_id, debit, credit, line_description, cost_center_id)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (entry_id, account_id_map[account_name], debit, credit, description, entry_cost_id)
            )
            line_added += 1

        if line_added == 0:
            cur.execute("DELETE FROM journal_entries WHERE id = ?", (entry_id,))
            continue

        entries_count += 1
        lines_count += line_added

    cur.execute(
        "INSERT INTO imports(source_file, imported_at, entries_count, lines_count) VALUES (?, ?, ?, ?)",
        (os.path.basename(file_path), datetime.now().isoformat(timespec="seconds"), entries_count, lines_count)
    )
    conn.commit()
    conn.close()

    result = (
        f"تم استيراد الملف: {os.path.basename(file_path)}\n"
        f"عدد الحسابات: {len(account_pairs)}\n"
        f"عدد القيود: {entries_count}\n"
        f"عدد الحركات: {lines_count}\n"
        "تم تحويل اليومية الأمريكية إلى قاعدة بيانات محاسبية جاهزة للتشغيل."
    )
    if cost_center_col:
        result += f"\nتم استيراد مراكز التكلفة تلقائياً: {len(center_ids)} مركز"
        if centers_created:
            result += "\n" + "، ".join(centers_created[:10]) + ("..." if len(centers_created) > 10 else "")
    return result
